# -*- coding: utf-8 -*-
# @Time ： 2023/8/3 16:38
# @Auth ： JeremyChim
# @File ：row_col.py
# @IDE ：PyCharm
# @Github https://github.com/JeremyChim/

import openpyxl as op

def ret_row_col(file,
                sheet:int):
    '''
    :param file: 文件的路径
    :param sheet: 表格的sheet号
    :return: 最大行数, 最大列数
    '''

    wb = op.load_workbook(file)
    sheet -= 1
    ws = wb.worksheets[sheet]

    row = ws.max_row
    col = ws.max_column
    # print('工作表列数：',ws.max_column)
    # print('工作表行数：',ws.max_row)

    return row, col

if __name__ == '__main__':
    file = r'D:\GithubProject\ZCP_Script\report.xlsx'
    sheet = 1

    row, col = ret_row_col(file, sheet)
    print(row.__class__, f'最大行数: {row}')
    print(col.__class__, f'最大列数: {col}')