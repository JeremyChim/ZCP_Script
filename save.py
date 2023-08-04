# -*- coding: utf-8 -*-
# @Time ： 2023/7/25 15:38
# @Auth ： JeremyChim
# @File ：write.py
# @IDE ：PyCharm
# @Github ：https://github.com/JeremyChim/

import openpyxl as op
from openpyxl.styles import Font, Alignment

def save_(file,
          sheet:int,
          cell_coord:str,
          cell_val:str,
          font_name:str = u'微软雅黑',
          font_size:int = 8,
          font_color:str = None,
          horz = None,
          vert = None):
    '''
    保存表格内容。

    :param file: 文件的路径
    :param sheet: 表格的sheet号
    :param cell_coord: 单元格的坐标，例如：A2
    :param cell_val: 写入的内容
    :param font_name: 字体
    :param font_size: 字号
    :param font_color: 字体颜色
    :param horz: 水平分布： center居中；left左；right右
    :param vert: 垂直分布： center居中；top上；bottom下
    '''

    wb = op.load_workbook(file)
    sheet -= 1
    ws = wb.worksheets[sheet]

    ws[cell_coord] = cell_val
    ws[cell_coord].font = Font(name=font_name, size=font_size, color=font_color)
    ws[cell_coord].alignment = Alignment(horizontal=horz, vertical=vert)

    wb.save(file)

if __name__ == '__main__':
    report = r'D:\GithubProject\ZCP_Script\report.xlsx'

    save_(file=report,
          sheet=1,
          cell_coord='A1',
          cell_val='COME ON. JER'
          )


